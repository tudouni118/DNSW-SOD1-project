import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import sys
import traceback
from datetime import datetime, date, timedelta
import calendar
import re

import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, Border, Side
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


def show_error(exc_type, exc_value, exc_tb):
    err_msg = ''.join(traceback.format_exception(exc_type, exc_value, exc_tb))
    print(err_msg, file=sys.stderr)
    try:
        messagebox.showerror("Error", f"Unhandled exception:\n\n{err_msg}")
    except:
        pass

sys.excepthook = show_error


def normalize_input(text):
    if not text:
        return text
    text = text.replace("。", ".")
    text = text.replace("．", ".")
    text = text.replace("、", ".")
    text = text.replace("，", ".")
    return text


def natural_sort_key(s):
    """自然排序：数字部分按数值排序"""
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r'(\d+)', str(s))]


class AutoCorrectEntry(ttk.Entry):
    def __init__(self, parent, textvariable=None, **kwargs):
        self._var = textvariable if textvariable else tk.StringVar()
        super().__init__(parent, textvariable=self._var, **kwargs)
        self.bind("<FocusOut>", self._on_focus_out)
        self.bind("<Return>", self._on_focus_out)

    def _on_focus_out(self, event=None):
        current = self._var.get()
        corrected = normalize_input(current)
        if corrected != current:
            self._var.set(corrected)


# ---------- 日历选择器 ----------
class DatePicker:
    def __init__(self, parent, target_var):
        self.parent = parent
        self.target_var = target_var
        self.today = date.today()
        self.selected_date = None
        self.current_month = self.today.month
        self.current_year = self.today.year
        self.day_buttons = {}

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Select Date")
        self.dialog.geometry("280x280")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.dialog.resizable(False, False)

        self._build()

    def _build(self):
        nav_frame = ttk.Frame(self.dialog)
        nav_frame.pack(fill="x", padx=5, pady=5)

        ttk.Button(nav_frame, text="◀", width=3, command=self._prev_month).pack(side=tk.LEFT)
        self.lbl_month = ttk.Label(nav_frame, text="", font=("", 10, "bold"))
        self.lbl_month.pack(side=tk.LEFT, expand=True)
        ttk.Button(nav_frame, text="▶", width=3, command=self._next_month).pack(side=tk.RIGHT)

        days_frame = ttk.Frame(self.dialog)
        days_frame.pack(fill="x", padx=5, pady=2)
        for d in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]:
            ttk.Label(days_frame, text=d, width=4, anchor="center", font=("", 8, "bold")).pack(side=tk.LEFT)

        self.grid_frame = ttk.Frame(self.dialog)
        self.grid_frame.pack(fill="both", expand=True, padx=5, pady=5)

        btn_frame = ttk.Frame(self.dialog)
        btn_frame.pack(fill="x", padx=5, pady=5)
        ttk.Button(btn_frame, text="Today", command=self._select_today).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=self.dialog.destroy).pack(side=tk.RIGHT, padx=5)

        self._refresh_calendar()

    def _get_button_color(self, d):
        if self.selected_date and d == self.selected_date:
            return "#0078d4"
        elif d == self.today:
            return "#cce5ff"
        else:
            return "#f0f0f0"

    def _refresh_calendar(self):
        for w in self.grid_frame.winfo_children():
            w.destroy()
        self.day_buttons.clear()

        self.lbl_month.config(text=f"{calendar.month_name[self.current_month]} {self.current_year}")
        cal = calendar.monthcalendar(self.current_year, self.current_month)

        for r, week in enumerate(cal):
            for c, day in enumerate(week):
                if day == 0:
                    ttk.Label(self.grid_frame, text="", width=4).grid(row=r, column=c, padx=1, pady=1)
                else:
                    d = date(self.current_year, self.current_month, day)
                    bg = self._get_button_color(d)

                    btn = tk.Button(
                        self.grid_frame, text=str(day), width=4,
                        bg=bg, relief="flat"
                    )
                    btn.bind("<Button-1>", lambda e, d=d: self._on_date_click(d))
                    btn.bind("<Double-Button-1>", lambda e, d=d: self._on_date_double_click(d))
                    btn.grid(row=r, column=c, padx=1, pady=1)
                    self.day_buttons[d] = btn

    def _update_button_colors(self):
        for d, btn in self.day_buttons.items():
            btn.config(bg=self._get_button_color(d))

    def _on_date_click(self, d):
        self.selected_date = d
        self._update_button_colors()

    def _on_date_double_click(self, d):
        self.target_var.set(d.strftime("%Y-%m-%d"))
        self.dialog.destroy()

    def _prev_month(self):
        if self.current_month == 1:
            self.current_month = 12
            self.current_year -= 1
        else:
            self.current_month -= 1
        self._refresh_calendar()

    def _next_month(self):
        if self.current_month == 12:
            self.current_month = 1
            self.current_year += 1
        else:
            self.current_month += 1
        self._refresh_calendar()

    def _select_today(self):
        self.target_var.set(self.today.strftime("%Y-%m-%d"))
        self.dialog.destroy()


class DateEntry(ttk.Frame):
    def __init__(self, parent, textvariable=None, **kwargs):
        super().__init__(parent)
        self.var = textvariable if textvariable else tk.StringVar()
        self.entry = ttk.Entry(self, textvariable=self.var, width=12, **kwargs)
        self.entry.pack(side=tk.LEFT)
        self.btn = ttk.Button(self, text="📅", width=3, command=self._open_picker)
        self.btn.pack(side=tk.LEFT, padx=2)

    def _open_picker(self):
        DatePicker(self, self.var)


NS_CRITERIA_TEXT = """NS=0 (Pre-symptomatic)
Upon tail suspension, the hindlimbs can splay normally, i.e., fully extended away from the lateral midline, and maintain this position for 2 seconds or longer; normal walking gait.
悬吊时后肢可以正常分开，即完全远离外侧中线，并在该位置停留2s或更长时间；可以正常行走。

NS=1 (First symptoms)
Upon tail suspension, the hindlimbs exhibit abnormal splaying, namely folding or partially folding towards the lateral midline, or displaying tremors, or the hindlimbs are in a retracted or clasped state; walking is normal or slow.
悬吊时后肢出现异常分开，即朝向外侧中线折叠或部分折叠，或出现震颤，或后肢处于收缩或交叉状态；行走正常或缓慢。

NS=2 (Mild paralysis)
Upon tail suspension, the hindlimbs are partially or completely folded with minimal extension (joint movements may still be present); during locomotion, the hindlimbs can advance but occasional dragging occurs; the animal can right itself within 10 seconds from either side.
悬吊时后肢部分或完全折叠，几乎不伸展（仍可能存在关节运动）；行走时后肢可以前进，但偶尔存在拖拽情况；双侧都可10秒内翻正。

NS=3 (Paralysis)
Upon tail suspension, the hindlimbs display rigid paralysis or minimal joint movement; forward locomotion is possible, but without the involvement of the hindlimbs; the animal can right itself within 10 seconds from either side.
悬吊时后肢出现僵硬瘫痪或极少关节运动；行走时可以前进，但后肢不参与；双侧都可10秒内翻正。

NS=4 (Humane endpoint)
Upon tail suspension, the hindlimbs display rigid paralysis; unable to move forward; failure to right itself from either side within 10-30 seconds.
悬吊时后肢出现僵硬性瘫痪；无法行进；任意一侧都无法10-30秒内翻正。"""


class DataManager:
    def __init__(self, filepath=None):
        if filepath is None:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            filepath = os.path.join(script_dir, "mouse_data.json")
        self.filepath = filepath
        self.data = {"mice": [], "birth_date": "2026-07-20"}
        self.load()
        self._migrate_old_data()

    def load(self):
        if os.path.exists(self.filepath):
            try:
                with open(self.filepath, "r", encoding="utf-8") as f:
                    self.data = json.load(f)
            except (json.JSONDecodeError, IOError):
                self.data = {"mice": [], "birth_date": "2026-07-20"}
        else:
            self.save()

    def save(self):
        with open(self.filepath, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False)

    def _migrate_old_data(self):
        if "birth_date" not in self.data:
            self.data["birth_date"] = "2026-07-20"
        for m in self.data["mice"]:
            if "tail_id" not in m and "id" in m:
                m["tail_id"] = m.pop("id")
            if "tail_id" not in m:
                m["tail_id"] = ""
            if "cage_id" not in m:
                m["cage_id"] = ""
            if "ear_tag" not in m:
                m["ear_tag"] = ""
            if "group" not in m:
                m["group"] = ""
            if "sub_group" not in m:
                m["sub_group"] = ""
            if "strain" not in m:
                m["strain"] = ""
            if "sex" not in m:
                m["sex"] = "Male"
            if "coat_color" not in m:
                m["coat_color"] = ""
            if "arrival_date" not in m:
                m["arrival_date"] = ""
            for r in m.get("records", []):
                if "grip" not in r:
                    r["grip"] = []
                if "rotarod" not in r:
                    r["rotarod"] = []
                if "balance_beam" not in r:
                    r["balance_beam"] = {"6mm": [], "12mm": []}
                if "rotarod_phase" not in r:
                    r["rotarod_phase"] = ""
                if "bb_phase" not in r:
                    r["bb_phase"] = ""
                if "neurological_score" not in r:
                    r["neurological_score"] = None
        self.save()

    def get_birth_date(self):
        return self.data.get("birth_date", "2026-07-20")

    def set_birth_date(self, date_str):
        self.data["birth_date"] = date_str
        self.save()

    def get_stats(self):
        total_mice = len(self.data["mice"])
        total_records = sum(len(m["records"]) for m in self.data["mice"])
        return total_mice, total_records

    def add_mouse(self, tail_id, cage_id="", group="", strain="", ear_tag="", sex="Male", sub_group="", coat_color="", arrival_date=""):
        if not tail_id:
            return False, "Mouse Tail ID is required!"
        for m in self.data["mice"]:
            existing_id = m.get("tail_id", m.get("id", ""))
            if existing_id == tail_id:
                return False, f"Mouse Tail ID '{tail_id}' already exists!"
        self.data["mice"].append({
            "tail_id": tail_id, "cage_id": cage_id, "ear_tag": ear_tag,
            "group": group, "sub_group": sub_group, "strain": strain, "sex": sex,
            "coat_color": coat_color, "arrival_date": arrival_date, "records": []
        })
        self.save()
        return True, f"Mouse '{tail_id}' added successfully."

    def update_mouse(self, tail_id, new_tail_id, cage_id, ear_tag, group, strain, sex, sub_group="", coat_color="", arrival_date=""):
        for m in self.data["mice"]:
            if m.get("tail_id", m.get("id", "")) == tail_id:
                if new_tail_id != tail_id:
                    for other in self.data["mice"]:
                        if other is m: continue
                        if other.get("tail_id", other.get("id", "")) == new_tail_id:
                            return False, f"Tail ID '{new_tail_id}' already exists!"
                m["tail_id"] = new_tail_id
                m["cage_id"] = cage_id
                m["ear_tag"] = ear_tag
                m["group"] = group
                m["sub_group"] = sub_group
                m["strain"] = strain
                m["sex"] = sex
                m["coat_color"] = coat_color
                m["arrival_date"] = arrival_date
                self.save()
                return True, "Updated successfully"
        return False, "Mouse not found"

    def batch_update_mice(self, tail_ids, updates):
        updated_count = 0
        for m in self.data["mice"]:
            if m.get("tail_id") in tail_ids:
                for key, value in updates.items():
                    if key in m:
                        m[key] = value
                updated_count += 1
        if updated_count > 0:
            self.save()
        return updated_count

    def regroup_mouse(self, tail_id, new_ear_tag, new_group, new_sub_group=""):
        for m in self.data["mice"]:
            if m.get("tail_id") == tail_id:
                m["ear_tag"] = new_ear_tag
                m["group"] = new_group
                m["sub_group"] = new_sub_group
                self.save()
                return True, f"Regrouped: {tail_id}"
        return False, "Mouse not found"

    def delete_mouse(self, tail_id):
        self.data["mice"] = [m for m in self.data["mice"] if m.get("tail_id") != tail_id]
        self.save()

    def delete_record(self, tail_id, date):
        for m in self.data["mice"]:
            if m.get("tail_id") == tail_id:
                m["records"] = [r for r in m["records"] if r["date"] != date]
                self.save()
                return True
        return False

    def add_record(self, tail_id, date, weight=None, grip=None, rotarod=None, balance_beam=None,
                   rotarod_phase="", bb_phase="", neurological_score=None):
        for m in self.data["mice"]:
            if m.get("tail_id") == tail_id:
                existing = None
                for r in m["records"]:
                    if r["date"] == date:
                        existing = r
                        break
                if existing:
                    if weight is not None: existing["weight"] = round(float(weight), 2)
                    if grip is not None and len(grip) > 0: existing["grip"] = [float(v) for v in grip]
                    if rotarod is not None and len(rotarod) > 0: existing["rotarod"] = rotarod
                    if rotarod_phase is not None: existing["rotarod_phase"] = rotarod_phase
                    if bb_phase is not None: existing["bb_phase"] = bb_phase
                    if neurological_score is not None: existing["neurological_score"] = neurological_score
                    if balance_beam is not None:
                        existing_bb = existing.get("balance_beam", {"6mm": [], "12mm": []})
                        for w in ["6mm", "12mm"]:
                            if balance_beam.get(w) and len(balance_beam[w]) > 0:
                                existing_bb[w] = [float(v) for v in balance_beam[w]]
                        existing["balance_beam"] = existing_bb
                    self.save()
                    return True, "Updated"
                else:
                    record = {
                        "date": date,
                        "weight": round(float(weight), 2) if weight is not None else None,
                        "grip": [float(v) for v in grip] if grip else [],
                        "rotarod": rotarod if rotarod else [],
                        "balance_beam": balance_beam if balance_beam else {"6mm": [], "12mm": []},
                        "rotarod_phase": rotarod_phase, "bb_phase": bb_phase,
                        "neurological_score": neurological_score
                    }
                    m["records"].append(record)
                    m["records"].sort(key=lambda x: x["date"])
                    self.save()
                    return True, "Recorded"
        return False, "Mouse not found"

    def get_tail_ids(self):
        return sorted([m.get("tail_id", m.get("id", "")) for m in self.data["mice"]], key=natural_sort_key)

    def get_ear_tags(self):
        tags = sorted(set(m.get("ear_tag", "") for m in self.data["mice"] if m.get("ear_tag", "")), key=natural_sort_key)
        return tags

    def get_mouse_info(self, tail_id):
        for m in self.data["mice"]:
            if m.get("tail_id") == tail_id: return m
        return None

    def get_mouse_by_ear_tag(self, ear_tag):
        for m in self.data["mice"]:
            if m.get("ear_tag", "") == ear_tag: return m
        return None

    def get_all_mice(self):
        return self.data["mice"]

    def backup(self):
        if os.path.exists(self.filepath):
            import shutil
            backup_path = self.filepath + ".backup"
            try:
                shutil.copy2(self.filepath, backup_path)
                return backup_path
            except: pass
        return None

    @staticmethod
    def compute_grip_max(grip_list):
        if grip_list and len(grip_list) > 0: return max(grip_list)
        return None

    @staticmethod
    def compute_rotarod_best(rotarod_list):
        if rotarod_list and len(rotarod_list) > 0:
            return max(rotarod_list, key=lambda x: x.get("latency_to_fall", 0))
        return None

    @staticmethod
    def compute_balance_beam_means(bb_data):
        result = {"6mm": None, "12mm": None}
        if bb_data:
            for w in ["6mm", "12mm"]:
                vals = bb_data.get(w, [])
                if vals and len(vals) > 0: result[w] = round(sum(vals) / len(vals), 2)
        return result

    @staticmethod
    def calc_age(birth_date_str, record_date_str):
        try:
            bd = datetime.strptime(birth_date_str, "%Y-%m-%d").date()
            rd = datetime.strptime(record_date_str, "%Y-%m-%d").date()
            return (rd - bd).days, round((rd - bd).days / 7, 2)
        except: return None, None


class MouseWeightApp:
    STRAIN_OPTIONS = ["WT", "B6SJL-SOD1G93A"]
    SEX_OPTIONS = ["Male", "Female"]
    GROUP_OPTIONS = ["", "1", "2", "3", "4", "5"]
    SUB_GROUP_OPTIONS = ["", "Main Study Group", "Satellite Group"]
    COAT_COLOR_OPTIONS = ["", "Brownish-grey / a", "Solid Black / A"]
    PHASE_OPTIONS = ["", "pre", "test1", "test2", "test3", "test4", "test5"]
    NS_OPTIONS = ["", "0", "1", "2", "3", "4"]

    ALL_COLUMNS = [
        ("Tail ID", "tail_id"), ("Cage", "cage_id"), ("Ear Tag", "ear_tag"),
        ("Group", "group"), ("Sub-Group", "sub_group"), ("Coat Color", "coat_color"),
        ("Strain", "strain"), ("Sex", "sex"), ("Arrival Date", "arrival_date"),
        ("Date", "date"), ("Wt(g)", "weight"),
        ("GripMax(g)", "grip_max"), ("RotaPhase", "rotarod_phase"),
        ("Rotarod(rpm/s)", "rotarod"), ("BBPhase", "bb_phase"),
        ("BB6mm(s)", "bb6mm"), ("BB12mm(s)", "bb12mm"), ("NS", "ns"),
    ]

    BATCH_EDITABLE_FIELDS = [
        ("Cage ID:", "cage_id", "entry"),
        ("Group:", "group", "group"),
        ("Sub-Group:", "sub_group", "sub_group"),
        ("Coat Color / Code:", "coat_color", "coat_color"),
        ("Strain:", "strain", "strain"),
        ("Sex:", "sex", "sex"),
        ("Arrival Date:", "arrival_date", "entry"),
    ]

    def __init__(self, root):
        self.root = root
        self.root.title("Mouse Weight & Behavior Manager")
        self.root.geometry("1400x850")
        self.dm = DataManager()
        self.current_fig = None
        self._last_selection_method = 'tail'

        self.column_visible = {}
        for col_name, _ in self.ALL_COLUMNS:
            self.column_visible[col_name] = tk.BooleanVar(value=True)

        self.sort_column = None
        self.sort_reverse = False

        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True, padx=5, pady=5)

        self.tab_input = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_input, text="Data Entry")
        self.setup_input_tab()

        self.tab_view = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_view, text="View Data")
        self.setup_view_tab()

        self.tab_chart = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_chart, text="Charts")
        self.setup_chart_tab()

        self.status = tk.Label(root, text="Ready", bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.status.pack(side=tk.BOTTOM, fill=tk.X)

        self.refresh_all()
        self.root.after(500, self.show_startup_info)

    def show_startup_info(self):
        total_mice, total_records = self.dm.get_stats()
        self.status.config(text=f"Data file: {self.dm.filepath} | {total_mice} mice, {total_records} records")
        messagebox.showinfo("Welcome", f"Data file: {self.dm.filepath}\n\nLoaded: {total_mice} mice, {total_records} records\nBirth date: {self.dm.get_birth_date()}")

    # ===================== 数据录入标签页 =====================
    def setup_input_tab(self):
        main_pw = tk.PanedWindow(self.tab_input, orient=tk.HORIZONTAL, sashrelief=tk.RAISED, sashwidth=4)
        main_pw.pack(fill="both", expand=True, padx=5, pady=5)

        # ---- 左列 ----
        left_frame = ttk.Frame(main_pw)
        main_pw.add(left_frame, minsize=280)

        add_frame = ttk.LabelFrame(left_frame, text="Add New Mouse")
        add_frame.pack(fill="x", padx=5, pady=5)

        fields = [
            ("Mouse Tail ID:", "entry_tail_id", "entry"),
            ("Cage ID:", "entry_cage", "entry"),
            ("Group:", "entry_group", "group"),
            ("Coat Color / Code:", "entry_coat_color", "coat_color"),
            ("Strain:", "entry_strain", "strain"),
            ("Sex:", "entry_sex", "sex"),
            ("Arrival Date:", "entry_arrival_date", "date"),
        ]
        self.input_vars = {}
        self.date_widgets = {}
        for i, (label, var_name, field_type) in enumerate(fields):
            ttk.Label(add_frame, text=label).grid(row=i, column=0, sticky="e", padx=5, pady=3)
            var = tk.StringVar()
            self.input_vars[var_name] = var
            if field_type == "strain":
                cb = ttk.Combobox(add_frame, textvariable=var, values=self.STRAIN_OPTIONS, state="readonly", width=18)
                cb.grid(row=i, column=1, sticky="w", padx=5, pady=3); cb.current(0)
            elif field_type == "sex":
                cb = ttk.Combobox(add_frame, textvariable=var, values=self.SEX_OPTIONS, state="readonly", width=18)
                cb.grid(row=i, column=1, sticky="w", padx=5, pady=3); cb.current(0)
            elif field_type == "group":
                cb = ttk.Combobox(add_frame, textvariable=var, values=self.GROUP_OPTIONS, state="readonly", width=18)
                cb.grid(row=i, column=1, sticky="w", padx=5, pady=3)
            elif field_type == "coat_color":
                cb = ttk.Combobox(add_frame, textvariable=var, values=self.COAT_COLOR_OPTIONS, state="readonly", width=18)
                cb.grid(row=i, column=1, sticky="w", padx=5, pady=3)
            elif field_type == "date":
                de = DateEntry(add_frame, textvariable=var)
                self.date_widgets[var_name] = de
                de.grid(row=i, column=1, sticky="w", padx=5, pady=3)
            else:
                AutoCorrectEntry(add_frame, textvariable=var, width=20).grid(row=i, column=1, sticky="w", padx=5, pady=3)

        ttk.Button(add_frame, text="Add Mouse", command=self.add_mouse).grid(row=7, column=0, columnspan=2, pady=10)

        bd_frame = ttk.Frame(add_frame)
        bd_frame.grid(row=8, column=0, columnspan=2, pady=5, sticky="w")
        ttk.Label(bd_frame, text="Birth Date:").pack(side=tk.LEFT, padx=5)
        self.var_birth_date = tk.StringVar(value=self.dm.get_birth_date())
        DateEntry(bd_frame, textvariable=self.var_birth_date).pack(side=tk.LEFT, padx=5)
        ttk.Button(bd_frame, text="Update", command=self.update_birth_date).pack(side=tk.LEFT, padx=5)

        ttk.Label(add_frame, text="Note: Only Tail ID is required.", foreground="gray", font=("", 8)).grid(row=9, column=0, columnspan=2, pady=5)

        # 重组区
        regroup_frame = ttk.LabelFrame(left_frame, text="Re-group Mouse")
        regroup_frame.pack(fill="x", padx=5, pady=5)

        ttk.Label(regroup_frame, text="Select Mouse:").grid(row=0, column=0, sticky="e", padx=5, pady=3)
        rg_frame = ttk.Frame(regroup_frame)
        rg_frame.grid(row=0, column=1, sticky="w", padx=5, pady=3)

        ttk.Label(rg_frame, text="Tail ID:", font=("", 8, "italic")).pack(side=tk.LEFT, padx=2)
        self.cb_regroup_mouse = ttk.Combobox(rg_frame, state="readonly", width=10)
        self.cb_regroup_mouse.pack(side=tk.LEFT, padx=2)
        self.cb_regroup_mouse.bind("<<ComboboxSelected>>", lambda e: self._on_regroup_tail_select())

        ttk.Label(rg_frame, text="Ear Tag:", font=("", 8, "italic")).pack(side=tk.LEFT, padx=(10, 2))
        self.cb_regroup_ear_tag = ttk.Combobox(rg_frame, state="readonly", width=10)
        self.cb_regroup_ear_tag.pack(side=tk.LEFT, padx=2)
        self.cb_regroup_ear_tag.bind("<<ComboboxSelected>>", lambda e: self._on_regroup_ear_select())

        ttk.Label(regroup_frame, text="Ear Tag:").grid(row=1, column=0, sticky="e", padx=5, pady=3)
        self.var_new_ear_tag = tk.StringVar()
        AutoCorrectEntry(regroup_frame, textvariable=self.var_new_ear_tag, width=20).grid(row=1, column=1, sticky="w", padx=5, pady=3)

        ttk.Label(regroup_frame, text="New Group ID:").grid(row=2, column=0, sticky="e", padx=5, pady=3)
        self.var_new_group = tk.StringVar()
        AutoCorrectEntry(regroup_frame, textvariable=self.var_new_group, width=20).grid(row=2, column=1, sticky="w", padx=5, pady=3)

        ttk.Label(regroup_frame, text="Sub-Group:").grid(row=3, column=0, sticky="e", padx=5, pady=3)
        self.var_new_sub_group = tk.StringVar()
        cb_sub = ttk.Combobox(regroup_frame, textvariable=self.var_new_sub_group, values=self.SUB_GROUP_OPTIONS, state="readonly", width=18)
        cb_sub.grid(row=3, column=1, sticky="w", padx=5, pady=3)

        ttk.Button(regroup_frame, text="Set", command=self.regroup_mouse).grid(row=4, column=0, columnspan=2, pady=10)

        # ---- 中列 ----
        middle_frame = ttk.Frame(main_pw)
        main_pw.add(middle_frame, minsize=450)

        middle_label = ttk.LabelFrame(middle_frame, text="Record Measurements")
        middle_label.pack(fill="both", expand=True)

        self._build_record_form(middle_label)

        # ---- 右列 ----
        right_frame = ttk.Frame(main_pw)
        main_pw.add(right_frame, minsize=380)

        ns_outer = ttk.LabelFrame(right_frame, text="Neurological Score (NS)")
        ns_outer.pack(fill="both", expand=True, padx=5, pady=5)

        ns_input_frame = ttk.Frame(ns_outer)
        ns_input_frame.pack(fill="x", padx=10, pady=10)
        ttk.Label(ns_input_frame, text="NS =", font=("", 11, "bold")).pack(side=tk.LEFT, padx=5)
        self.var_ns = tk.StringVar()
        cb_ns = ttk.Combobox(ns_input_frame, textvariable=self.var_ns, values=self.NS_OPTIONS, state="readonly", width=5, font=("", 11))
        cb_ns.pack(side=tk.LEFT, padx=5)
        ttk.Label(ns_input_frame, text="(select or leave blank)", foreground="gray", font=("", 9)).pack(side=tk.LEFT, padx=10)

        ttk.Separator(ns_outer, orient="horizontal").pack(fill="x", padx=10, pady=5)
        ttk.Label(ns_outer, text="Scoring Criteria:", font=("", 10, "bold")).pack(anchor="w", padx=10, pady=5)

        ns_text_frame = ttk.Frame(ns_outer)
        ns_text_frame.pack(fill="x", padx=10, pady=5)
        ns_text = tk.Text(ns_text_frame, wrap=tk.WORD, font=("", 8), height=20)
        ns_text.insert("1.0", NS_CRITERIA_TEXT)
        ns_text.configure(state="disabled")
        ns_scroll = ttk.Scrollbar(ns_text_frame, orient="vertical", command=ns_text.yview)
        ns_text.configure(yscrollcommand=ns_scroll.set)
        ns_text.pack(side=tk.LEFT, fill="x", expand=True)
        ns_scroll.pack(side=tk.RIGHT, fill="y")

        ttk.Label(ns_outer, text="[Reserved for future functions]", foreground="lightgray", font=("", 9, "italic")).pack(pady=20)

    def _build_record_form(self, parent):
        row = 0

        ttk.Label(parent, text="Select Mouse:", font=("", 10, "bold")).grid(row=row, column=0, sticky="w", padx=5, pady=3)
        search_frame = ttk.Frame(parent)
        search_frame.grid(row=row, column=1, sticky="w", padx=5, pady=3)

        ttk.Label(search_frame, text="Tail ID:", font=("", 8, "italic")).pack(side=tk.LEFT, padx=2)
        self.cb_mouse = ttk.Combobox(search_frame, state="readonly", width=10)
        self.cb_mouse.pack(side=tk.LEFT, padx=2)
        self.cb_mouse.bind("<<ComboboxSelected>>", lambda e: self.on_mouse_select())

        ttk.Label(search_frame, text="Ear Tag:", font=("", 8, "italic")).pack(side=tk.LEFT, padx=(10, 2))
        self.cb_mouse_ear = ttk.Combobox(search_frame, state="readonly", width=10)
        self.cb_mouse_ear.pack(side=tk.LEFT, padx=2)
        self.cb_mouse_ear.bind("<<ComboboxSelected>>", lambda e: self._on_ear_tag_select())

        row += 1

        ttk.Label(parent, text="Date (YYYY-MM-DD):").grid(row=row, column=0, sticky="w", padx=5, pady=3)
        date_frame = ttk.Frame(parent)
        date_frame.grid(row=row, column=1, sticky="w", padx=5, pady=3)
        self.var_date = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.var_date.trace_add("write", lambda *a: self.update_age_display())
        DateEntry(date_frame, textvariable=self.var_date).pack(side=tk.LEFT)
        self.lbl_wks = ttk.Label(date_frame, text="", foreground="blue")
        self.lbl_wks.pack(side=tk.LEFT, padx=10)
        self.lbl_days = ttk.Label(date_frame, text="", foreground="green")
        self.lbl_days.pack(side=tk.LEFT, padx=5)
        self.update_age_display()
        row += 1

        ttk.Separator(parent, orient="horizontal").grid(row=row, column=0, columnspan=2, sticky="ew", pady=5)
        row += 1

        # 体重
        weight_frame = ttk.LabelFrame(parent, text="Body Weight (g)")
        weight_frame.grid(row=row, column=0, columnspan=2, sticky="ew", padx=5, pady=3)
        ttk.Label(weight_frame, text="Weight:").grid(row=0, column=0, padx=5, pady=3)
        self.var_weight = tk.StringVar()
        self.entry_weight = AutoCorrectEntry(weight_frame, textvariable=self.var_weight, width=12)
        self.entry_weight.grid(row=0, column=1, padx=5, pady=3)
        self.entry_weight.bind("<FocusOut>", self.format_weight)
        row += 1

        # 抓力
        grip_frame = ttk.LabelFrame(parent, text="Grip Strength (g) — 10 values, max used")
        grip_frame.grid(row=row, column=0, columnspan=2, sticky="ew", padx=5, pady=3)
        self.grip_vars = []
        for i in range(5):
            var = tk.StringVar(); self.grip_vars.append(var)
            ttk.Label(grip_frame, text=f"{i+1}:", font=("", 9)).grid(row=0, column=i*2, padx=1, pady=1, sticky="e")
            AutoCorrectEntry(grip_frame, textvariable=var, width=6).grid(row=0, column=i*2+1, padx=1, pady=1)
        for i in range(5, 10):
            var = tk.StringVar(); self.grip_vars.append(var)
            ttk.Label(grip_frame, text=f"{i+1}:", font=("", 9)).grid(row=1, column=(i-5)*2, padx=1, pady=1, sticky="e")
            AutoCorrectEntry(grip_frame, textvariable=var, width=6).grid(row=1, column=(i-5)*2+1, padx=1, pady=1)
        btn_row = ttk.Frame(grip_frame)
        btn_row.grid(row=2, column=0, columnspan=10, pady=3)
        ttk.Button(btn_row, text="Compute Max", command=self.compute_grip_max_display).pack(side=tk.LEFT, padx=5)
        self.lbl_grip_max = ttk.Label(btn_row, text="Max: --", foreground="red")
        self.lbl_grip_max.pack(side=tk.LEFT, padx=10)
        row += 1

        # Rotarod
        rotarod_frame = ttk.LabelFrame(parent, text="Rotarod — 3 trials")
        rotarod_frame.grid(row=row, column=0, columnspan=2, sticky="ew", padx=5, pady=3)
        rota_phase_frame = ttk.Frame(rotarod_frame)
        rota_phase_frame.grid(row=0, column=0, columnspan=7, sticky="w", padx=5, pady=2)
        ttk.Label(rota_phase_frame, text="Phase:").pack(side=tk.LEFT, padx=2)
        self.var_rotarod_phase = tk.StringVar(value="")
        cb_rota = ttk.Combobox(rota_phase_frame, textvariable=self.var_rotarod_phase, values=self.PHASE_OPTIONS, state="readonly", width=8)
        cb_rota.pack(side=tk.LEFT, padx=2)

        self.rotarod_vars = []
        for i in range(3):
            trial_frame = ttk.Frame(rotarod_frame)
            trial_frame.grid(row=i+1, column=0, columnspan=7, sticky="w", padx=5, pady=2)
            ttk.Label(trial_frame, text=f"Trial {i+1}:").pack(side=tk.LEFT, padx=2)
            ttk.Label(trial_frame, text="Speed:").pack(side=tk.LEFT, padx=2)
            var_speed = tk.StringVar(); AutoCorrectEntry(trial_frame, textvariable=var_speed, width=7).pack(side=tk.LEFT, padx=2)
            ttk.Label(trial_frame, text="rpm").pack(side=tk.LEFT, padx=2)
            ttk.Label(trial_frame, text="Latency to fall:").pack(side=tk.LEFT, padx=2)
            var_latency = tk.StringVar(); AutoCorrectEntry(trial_frame, textvariable=var_latency, width=7).pack(side=tk.LEFT, padx=2)
            ttk.Label(trial_frame, text="s").pack(side=tk.LEFT, padx=2)
            self.rotarod_vars.append((var_speed, var_latency))
        btn_best_frame = ttk.Frame(rotarod_frame)
        btn_best_frame.grid(row=4, column=0, columnspan=7, pady=3)
        ttk.Button(btn_best_frame, text="Find Best", command=self.compute_rotarod_best_display).pack(side=tk.LEFT, padx=5)
        self.lbl_rotarod_best = ttk.Label(btn_best_frame, text="Best: --", foreground="red")
        self.lbl_rotarod_best.pack(side=tk.LEFT, padx=10)
        row += 1

        # Balance Beam
        bb_frame = ttk.LabelFrame(parent, text="Balance Beam — 2 widths × 2 trials")
        bb_frame.grid(row=row, column=0, columnspan=2, sticky="ew", padx=5, pady=3)
        bb_phase_frame = ttk.Frame(bb_frame)
        bb_phase_frame.grid(row=0, column=0, columnspan=10, sticky="w", padx=5, pady=2)
        ttk.Label(bb_phase_frame, text="Phase:").pack(side=tk.LEFT, padx=2)
        self.var_bb_phase = tk.StringVar(value="")
        cb_bb = ttk.Combobox(bb_phase_frame, textvariable=self.var_bb_phase, values=self.PHASE_OPTIONS, state="readonly", width=8)
        cb_bb.pack(side=tk.LEFT, padx=2)

        self.bb_vars = {}
        bb_row = 1
        for width in ["6mm", "12mm"]:
            width_frame = ttk.Frame(bb_frame)
            width_frame.grid(row=bb_row, column=0, columnspan=10, sticky="w", padx=5, pady=3)
            ttk.Label(width_frame, text=f"{width}:", font=("", 9, "bold")).pack(side=tk.LEFT, padx=5)
            bb_vars_width = []
            for ti in range(2):
                ttk.Label(width_frame, text=f"Trial {ti+1}: cross time:").pack(side=tk.LEFT, padx=2)
                var = tk.StringVar(); AutoCorrectEntry(width_frame, textvariable=var, width=7).pack(side=tk.LEFT, padx=2)
                ttk.Label(width_frame, text="s").pack(side=tk.LEFT, padx=5)
                bb_vars_width.append(var)
            self.bb_vars[width] = bb_vars_width
            bb_row += 1
        btn_means_frame = ttk.Frame(bb_frame)
        btn_means_frame.grid(row=bb_row, column=0, columnspan=10, pady=3)
        ttk.Button(btn_means_frame, text="Compute Means", command=self.compute_bb_means_display).pack(side=tk.LEFT, padx=5)
        self.lbl_bb_means = ttk.Label(btn_means_frame, text="Means: 6mm=--, 12mm=--", foreground="red")
        self.lbl_bb_means.pack(side=tk.LEFT, padx=10)
        row += 1

        self.lbl_info = ttk.Label(parent, text="", foreground="gray")
        self.lbl_info.grid(row=row, column=0, columnspan=2, pady=3)
        row += 1

        # Save
        ttk.Separator(parent, orient="horizontal").grid(row=row, column=0, columnspan=2, sticky="ew", pady=5)
        row += 1
        save_frame = ttk.Frame(parent)
        save_frame.grid(row=row, column=0, columnspan=2, pady=10)
        ttk.Button(save_frame, text="💾 Save Record", command=self.add_record).pack(side=tk.LEFT, padx=5)
        ttk.Label(save_frame, text="(only filled fields will be saved/updated)", foreground="gray", font=("", 8)).pack(side=tk.LEFT, padx=10)

    def _on_regroup_tail_select(self):
        tail_id = self.cb_regroup_mouse.get()
        mouse = self.dm.get_mouse_info(tail_id)
        if mouse:
            self.cb_regroup_ear_tag.set(mouse.get("ear_tag", ""))

    def _on_regroup_ear_select(self):
        ear_tag = self.cb_regroup_ear_tag.get()
        mouse = self.dm.get_mouse_by_ear_tag(ear_tag)
        if mouse:
            self.cb_regroup_mouse.set(mouse.get("tail_id", ""))

    def _on_ear_tag_select(self):
        self._last_selection_method = 'ear'
        ear_tag = self.cb_mouse_ear.get()
        mouse = self.dm.get_mouse_by_ear_tag(ear_tag)
        if mouse:
            self.cb_mouse.set(mouse.get("tail_id", ""))
            self.on_mouse_select()

    def update_age_display(self):
        date_str = self.var_date.get().strip()
        birth_str = self.dm.get_birth_date()
        days, wks = self.dm.calc_age(birth_str, date_str)
        if days is not None:
            self.lbl_days.config(text=f"{days} Days")
            self.lbl_wks.config(text=f"{wks:.2f} wks")
        else:
            self.lbl_days.config(text=""); self.lbl_wks.config(text="")

    def format_weight(self, event=None):
        val = normalize_input(self.var_weight.get().strip())
        if not val:
            return
        try:
            num = float(val)
            if '.' in val:
                decimals = len(val.split('.')[1])
                if decimals > 2:
                    messagebox.showerror("Error", "Weight must have at most 2 decimal places!")
                    self.var_weight.set("")
                    self.entry_weight.focus_set()
                    return
            self.var_weight.set(f"{num:.2f}")
        except ValueError:
            pass

    def compute_grip_max_display(self):
        vals = []
        for v in self.grip_vars:
            txt = normalize_input(v.get().strip())
            if txt:
                try: vals.append(float(txt))
                except ValueError: pass
        self.lbl_grip_max.config(text=f"Max: {max(vals):.2f} g" if vals else "Max: --")

    def compute_rotarod_best_display(self):
        trials = []
        for vs, vl in self.rotarod_vars:
            s = normalize_input(vs.get().strip()); l = normalize_input(vl.get().strip())
            if s and l:
                try: trials.append({"speed": float(s), "latency_to_fall": float(l)})
                except ValueError: pass
        if trials:
            best = max(trials, key=lambda x: x["latency_to_fall"])
            self.lbl_rotarod_best.config(text=f"Best: Speed={best['speed']}rpm, Latency to fall={best['latency_to_fall']}s")
        else:
            self.lbl_rotarod_best.config(text="Best: --")

    def compute_bb_means_display(self):
        means = {}
        for width in ["6mm", "12mm"]:
            vals = []
            for v in self.bb_vars[width]:
                txt = normalize_input(v.get().strip())
                if txt:
                    try: vals.append(float(txt))
                    except ValueError: pass
            means[width] = round(sum(vals)/len(vals), 2) if vals else None
        self.lbl_bb_means.config(text=f"Means: 6mm={means['6mm']}s, 12mm={means['12mm']}s" if means['6mm'] is not None or means['12mm'] is not None else "Means: --")

    def update_birth_date(self):
        new_date = normalize_input(self.var_birth_date.get().strip())
        try:
            datetime.strptime(new_date, "%Y-%m-%d")
            self.dm.set_birth_date(new_date); self.update_age_display()
            messagebox.showinfo("Success", f"Birth date updated to {new_date}")
        except ValueError:
            messagebox.showerror("Error", "Invalid date format.")

    def add_mouse(self):
        tail_id = normalize_input(self.input_vars["entry_tail_id"].get().strip())
        cage = normalize_input(self.input_vars["entry_cage"].get().strip())
        group = self.input_vars["entry_group"].get().strip()
        coat_color = self.input_vars["entry_coat_color"].get().strip()
        strain = self.input_vars["entry_strain"].get().strip()
        sex = self.input_vars["entry_sex"].get().strip()
        arrival_date = normalize_input(self.input_vars["entry_arrival_date"].get().strip())
        if not tail_id:
            messagebox.showwarning("Warning", "Mouse Tail ID is required.")
            return
        success, msg = self.dm.add_mouse(tail_id, cage, group, strain, "", sex if sex else "Male", "", coat_color, arrival_date)
        if success:
            messagebox.showinfo("Success", msg)
            self.input_vars["entry_tail_id"].set("")
            self.refresh_all()
        else:
            messagebox.showerror("Duplicate Entry", msg)

    def regroup_mouse(self):
        tail_id = self.cb_regroup_mouse.get()
        new_ear_tag = normalize_input(self.var_new_ear_tag.get().strip())
        new_group = normalize_input(self.var_new_group.get().strip())
        new_sub_group = self.var_new_sub_group.get().strip()
        if not tail_id:
            messagebox.showwarning("Warning", "Please select a mouse"); return
        mouse = self.dm.get_mouse_info(tail_id)
        if mouse:
            if not new_ear_tag: new_ear_tag = mouse.get("ear_tag", "")
            if not new_group: new_group = mouse.get("group", "")
            if not new_sub_group: new_sub_group = mouse.get("sub_group", "")
        success, msg = self.dm.regroup_mouse(tail_id, new_ear_tag, new_group, new_sub_group)
        if success: messagebox.showinfo("Success", msg); self.refresh_all()
        else: messagebox.showerror("Error", msg)

    def add_record(self):
        mid = self.cb_mouse.get()
        date = normalize_input(self.var_date.get().strip())
        if not mid or not date:
            messagebox.showwarning("Warning", "Please select a mouse and enter a date"); return

        weight = None
        weight_str = normalize_input(self.var_weight.get().strip())
        if weight_str:
            try:
                weight = float(weight_str)
                if '.' in weight_str:
                    decimals = len(weight_str.split('.')[1])
                    if decimals > 2:
                        messagebox.showerror("Error", "Weight must have at most 2 decimal places!")
                        return
                weight = round(weight, 2)
            except ValueError: messagebox.showerror("Error", "Weight must be a number"); return

        grip = []
        for v in self.grip_vars:
            txt = normalize_input(v.get().strip())
            if txt:
                try: grip.append(float(txt))
                except ValueError: messagebox.showerror("Error", f"Grip value '{txt}' invalid"); return

        rotarod = []
        for vs, vl in self.rotarod_vars:
            s = normalize_input(vs.get().strip()); l = normalize_input(vl.get().strip())
            if s and l:
                try: rotarod.append({"speed": float(s), "latency_to_fall": float(l)})
                except ValueError: messagebox.showerror("Error", "Rotarod values must be numbers"); return
            elif s or l:
                messagebox.showerror("Error", "Rotarod: each trial needs both values"); return

        balance_beam = {"6mm": [], "12mm": []}
        for width in ["6mm", "12mm"]:
            for v in self.bb_vars[width]:
                txt = normalize_input(v.get().strip())
                if txt:
                    try: balance_beam[width].append(float(txt))
                    except ValueError: messagebox.showerror("Error", f"BB {width} invalid"); return

        rotarod_phase = self.var_rotarod_phase.get().strip()
        bb_phase = self.var_bb_phase.get().strip()

        ns_str = self.var_ns.get().strip()
        neurological_score = None
        if ns_str:
            try:
                ns_val = int(ns_str)
                if ns_val < 0 or ns_val > 4: messagebox.showerror("Error", "NS must be 0-4"); return
                neurological_score = ns_val
            except ValueError: messagebox.showerror("Error", "NS must be integer"); return

        has_data = (weight is not None or len(grip) > 0 or len(rotarod) > 0 or
                    len(balance_beam["6mm"]) > 0 or len(balance_beam["12mm"]) > 0 or neurological_score is not None)
        if not has_data:
            messagebox.showwarning("Warning", "No data entered."); return

        success, msg = self.dm.add_record(mid, date, weight, grip, rotarod, balance_beam, rotarod_phase, bb_phase, neurological_score)
        if success:
            self.status.config(text=f"{msg} - {mid} @ {date}")
            # 保存前先记录用户当前选择的“方式”和“编号”，供保存后自动预填下一只
            selection_method = self._last_selection_method
            current_value = self.cb_mouse_ear.get() if selection_method == 'ear' else mid
            self.clear_record_form()
            self.refresh_all()
            self._select_next_mouse(selection_method, current_value)
        else:
            messagebox.showerror("Error", msg)

    def _select_next_mouse(self, selection_method=None, current_value=None):
        """保存后自动切换到下一只老鼠（尾标或耳标，取决于用户选择方式）。

        尾标/耳标均按自然排序（数字部分按数值）顺序，保存后自动预填下一个编号；
        若已是最后一个，则回到第一个。
        """
        if selection_method is None:
            selection_method = self._last_selection_method
        if selection_method == 'ear':
            ear_tags = self.dm.get_ear_tags()
            if current_value is None:
                current_value = self.cb_mouse_ear.get()
            if current_value in ear_tags:
                idx = ear_tags.index(current_value)
                next_ear = ear_tags[(idx + 1) % len(ear_tags)] if ear_tags else ""
                if next_ear:
                    self.cb_mouse_ear.set(next_ear)
                    self._on_ear_tag_select()
        else:
            tail_ids = self.dm.get_tail_ids()
            if current_value is None:
                current_value = self.cb_mouse.get()
            if current_value in tail_ids:
                idx = tail_ids.index(current_value)
                next_tail = tail_ids[(idx + 1) % len(tail_ids)] if tail_ids else ""
                if next_tail:
                    self.cb_mouse.set(next_tail)
                    self.on_mouse_select()

    def on_mouse_select(self):
        self._last_selection_method = 'tail'
        mid = self.cb_mouse.get()
        mouse = self.dm.get_mouse_info(mid)
        if mouse:
            self.cb_mouse_ear.set(mouse.get("ear_tag", ""))
            recs = mouse["records"]
            last = f"Latest: {recs[-1]['date']} | Wt:{recs[-1].get('weight','--')}g" if recs else "No records"
            self.lbl_info.config(text=f"Cage:{mouse['cage_id']} | EarTag:{mouse['ear_tag']} | Group:{mouse['group']} | {last}")

    def clear_record_form(self):
        self.var_weight.set("")
        for v in self.grip_vars: v.set("")
        for vs, vl in self.rotarod_vars: vs.set(""); vl.set("")
        for width in ["6mm", "12mm"]:
            for v in self.bb_vars[width]: v.set("")
        self.var_ns.set("")
        self.lbl_grip_max.config(text="Max: --"); self.lbl_rotarod_best.config(text="Best: --")
        self.lbl_bb_means.config(text="Means: 6mm=--, 12mm=--")

    # ===================== 数据查看标签页 =====================
    def setup_view_tab(self):
        top_frame = ttk.Frame(self.tab_view)
        top_frame.pack(fill="x", padx=10, pady=5)

        filter_frame = ttk.LabelFrame(top_frame, text="Filter")
        filter_frame.pack(side=tk.LEFT, fill="x", padx=5, pady=5)

        ttk.Label(filter_frame, text="Group:").grid(row=0, column=0, padx=5, pady=2)
        self.cb_filter_group = ttk.Combobox(filter_frame, state="readonly", width=15)
        self.cb_filter_group.grid(row=0, column=1, padx=5, pady=2)
        self.cb_filter_group.bind("<<ComboboxSelected>>", lambda e: self.refresh_view())

        ttk.Label(filter_frame, text="Select Mice:").grid(row=1, column=0, padx=5, pady=2)
        sel_frame = ttk.Frame(filter_frame)
        sel_frame.grid(row=1, column=1, padx=5, pady=2)

        ttk.Label(sel_frame, text="Tail:", font=("", 8, "italic")).pack(side=tk.LEFT, padx=2)
        self.listbox_filter_mice = tk.Listbox(sel_frame, selectmode=tk.MULTIPLE, height=4, width=12, exportselection=False)
        self.listbox_filter_mice.pack(side=tk.LEFT, padx=2)
        self.listbox_filter_mice.bind("<<ListboxSelect>>", lambda e: self._on_filter_tail_select())

        ttk.Label(sel_frame, text="Ear:", font=("", 8, "italic")).pack(side=tk.LEFT, padx=(10, 2))
        self.listbox_filter_ear = tk.Listbox(sel_frame, selectmode=tk.MULTIPLE, height=4, width=12, exportselection=False)
        self.listbox_filter_ear.pack(side=tk.LEFT, padx=2)
        self.listbox_filter_ear.bind("<<ListboxSelect>>", lambda e: self._on_filter_ear_select())

        ttk.Button(filter_frame, text="Show All", command=self.reset_filter).grid(row=2, column=0, columnspan=2, pady=5)

        column_frame = ttk.LabelFrame(top_frame, text="Show/Hide Columns")
        column_frame.pack(side=tk.LEFT, fill="x", padx=5, pady=5, expand=True)

        self.col_checkbuttons = {}
        rows_needed = (len(self.ALL_COLUMNS) + 5) // 6
        for i, (col_name, _) in enumerate(self.ALL_COLUMNS):
            r = i // 6; c = i % 6
            cb = ttk.Checkbutton(column_frame, text=col_name, variable=self.column_visible[col_name], command=self.refresh_view)
            cb.grid(row=r, column=c, sticky="w", padx=3, pady=2)
            self.col_checkbuttons[col_name] = cb

        ttk.Button(column_frame, text="Select All", command=self.select_all_columns).grid(row=rows_needed, column=0, padx=5, pady=3)
        ttk.Button(column_frame, text="Deselect All", command=self.deselect_all_columns).grid(row=rows_needed, column=1, padx=5, pady=3)

        export_frame = ttk.LabelFrame(top_frame, text="Export")
        export_frame.pack(side=tk.RIGHT, padx=5, pady=5)
        ttk.Button(export_frame, text="📊 Export to Excel", command=self.export_to_excel).pack(padx=10, pady=10)
        if not HAS_EXCEL:
            ttk.Label(export_frame, text="(pip install openpyxl)", foreground="gray", font=("", 8)).pack(padx=5, pady=2)

        self.tree = ttk.Treeview(self.tab_view, show="headings", selectmode="extended")
        self.tree.pack(fill="both", expand=True, padx=10, pady=5)

        scrollbar_y = ttk.Scrollbar(self.tree, orient="vertical", command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(self.tree, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)

        self.tree.bind("<Double-1>", self.on_tree_double_click)

        btn_frame = ttk.Frame(self.tab_view)
        btn_frame.pack(fill="x", padx=10, pady=5)
        ttk.Button(btn_frame, text="Delete Selected Mouse", command=self.delete_selected_mouse).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Delete Selected Record", command=self.delete_selected_record).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="✏️ Batch Edit Mice", command=self.batch_edit_mice).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Create Backup", command=self.create_backup).pack(side=tk.LEFT, padx=5)

    def _on_filter_tail_select(self):
        self.listbox_filter_ear.selection_clear(0, tk.END)
        self.refresh_view()

    def _on_filter_ear_select(self):
        self.listbox_filter_mice.selection_clear(0, tk.END)
        self.refresh_view()

    def select_all_columns(self):
        for col_name in self.column_visible: self.column_visible[col_name].set(True)
        self.refresh_view()

    def deselect_all_columns(self):
        for col_name in self.column_visible: self.column_visible[col_name].set(False)
        self.refresh_view()

    def reset_filter(self):
        self.cb_filter_group.set("")
        self.listbox_filter_mice.selection_clear(0, tk.END)
        self.listbox_filter_ear.selection_clear(0, tk.END)
        self.refresh_view()

    def refresh_view(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        visible_col_names = []
        visible_col_keys = []
        for col_name, col_key in self.ALL_COLUMNS:
            if self.column_visible[col_name].get():
                visible_col_names.append(col_name)
                visible_col_keys.append(col_key)

        if not visible_col_names:
            visible_col_names = ["Tail ID"]
            visible_col_keys = ["tail_id"]

        self.tree.configure(columns=visible_col_names, displaycolumns=visible_col_names)
        for col_name in visible_col_names:
            self.tree.heading(col_name, text=col_name, command=lambda c=col_name: self.sort_by_column(c))
            self.tree.column(col_name, width=90, minwidth=60)

        filter_group = self.cb_filter_group.get()
        selected_tails = [self.listbox_filter_mice.get(i) for i in self.listbox_filter_mice.curselection()]
        selected_ears = [self.listbox_filter_ear.get(i) for i in self.listbox_filter_ear.curselection()]

        rows = []
        for m in self.dm.get_all_mice():
            tail_id = m.get("tail_id", "")
            if filter_group and m["group"] != filter_group: continue
            if selected_tails and tail_id not in selected_tails: continue
            if selected_ears and m.get("ear_tag", "") not in selected_ears: continue
            if not m["records"]:
                rows.append(self._build_row_data(m, None, visible_col_keys))
            for r in m["records"]:
                rows.append(self._build_row_data(m, r, visible_col_keys))

        if self.sort_column and self.sort_column in visible_col_names:
            col_idx = visible_col_names.index(self.sort_column)
            try: rows.sort(key=lambda x: self._sort_key(x[col_idx]), reverse=self.sort_reverse)
            except: pass

        for row_data in rows:
            self.tree.insert("", "end", values=row_data)

    def _sort_key(self, value):
        if value is None or value == "": return (0, "")
        if isinstance(value, (int, float)): return (1, value)
        try: return (1, float(str(value).replace(",", ".")))
        except: return (2, str(value).lower())

    def sort_by_column(self, col_name):
        if self.sort_column == col_name:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col_name; self.sort_reverse = False
        self.refresh_view()

    def _build_row_data(self, mouse, record, visible_col_keys):
        values = []
        for key in visible_col_keys:
            if key == "tail_id": values.append(mouse.get("tail_id", ""))
            elif key == "cage_id": values.append(mouse.get("cage_id", ""))
            elif key == "ear_tag": values.append(mouse.get("ear_tag", ""))
            elif key == "group": values.append(mouse.get("group", ""))
            elif key == "sub_group": values.append(mouse.get("sub_group", ""))
            elif key == "coat_color": values.append(mouse.get("coat_color", ""))
            elif key == "strain": values.append(mouse.get("strain", ""))
            elif key == "sex": values.append(mouse.get("sex", ""))
            elif key == "arrival_date": values.append(mouse.get("arrival_date", ""))
            elif key == "date": values.append(record["date"] if record else "No records")
            elif key == "weight":
                val = record.get("weight") if record else None
                values.append(f"{val:.2f}" if val is not None else "")
            elif key == "grip_max":
                if record:
                    gm = self.dm.compute_grip_max(record.get("grip", []))
                    values.append(f"{gm:.2f}" if gm is not None else "")
                else: values.append("")
            elif key == "rotarod_phase": values.append(record.get("rotarod_phase", "") if record else "")
            elif key == "rotarod":
                if record:
                    best = self.dm.compute_rotarod_best(record.get("rotarod", []))
                    values.append(f"{best['speed']}/{best['latency_to_fall']}" if best else "")
                else: values.append("")
            elif key == "bb_phase": values.append(record.get("bb_phase", "") if record else "")
            elif key == "bb6mm":
                if record:
                    m_val = self.dm.compute_balance_beam_means(record.get("balance_beam", {}))
                    val = m_val.get("6mm")
                    values.append(f"{val:.2f}" if val is not None else "")
                else: values.append("")
            elif key == "bb12mm":
                if record:
                    m_val = self.dm.compute_balance_beam_means(record.get("balance_beam", {}))
                    val = m_val.get("12mm")
                    values.append(f"{val:.2f}" if val is not None else "")
                else: values.append("")
            elif key == "ns":
                ns = record.get("neurological_score") if record else None
                values.append(str(ns) if ns is not None else "")
            else: values.append("")
        return values

    def on_tree_double_click(self, event):
        sel = self.tree.selection()
        if not sel: return
        vals = self.tree.item(sel[0], "values")
        if not vals: return
        cols = list(self.tree["columns"])
        tail_idx = cols.index("Tail ID") if "Tail ID" in cols else 0
        date_idx = cols.index("Date") if "Date" in cols else 1
        tail_id = vals[tail_idx]
        date_val = vals[date_idx] if date_idx < len(vals) else ""
        mouse = self.dm.get_mouse_info(tail_id)
        if not mouse: return

        record = None
        if date_val and date_val != "No records":
            for r in mouse["records"]:
                if r["date"] == date_val:
                    record = r
                    break

        self.edit_combined_dialog(mouse, record, date_val if date_val != "No records" else "")

    def edit_combined_dialog(self, mouse, record, record_date):
        dialog = tk.Toplevel(self.root)
        tail_id = mouse.get("tail_id", "")
        title = f"Edit: {tail_id}"
        if record_date:
            title += f" @ {record_date}"
        dialog.title(title)
        dialog.geometry("550x750")
        dialog.transient(self.root)
        dialog.grab_set()

        canvas = tk.Canvas(dialog, width=530, highlightthickness=0)
        scrollbar = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        frame = ttk.Frame(canvas)
        frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill="both", expand=True)
        scrollbar.pack(side=tk.RIGHT, fill="y")

        row = 0

        ttk.Label(frame, text="Basic Information", font=("", 11, "bold")).grid(
            row=row, column=0, columnspan=2, sticky="w", padx=10, pady=8)
        row += 1

        basic_fields = [
            ("Tail ID:", "tail_id", "entry"),
            ("Cage ID:", "cage_id", "entry"),
            ("Ear Tag:", "ear_tag", "entry"),
            ("Group:", "group", "group"),
            ("Sub-Group:", "sub_group", "sub_group"),
            ("Coat Color / Code:", "coat_color", "coat_color"),
            ("Strain:", "strain", "strain"),
            ("Sex:", "sex", "sex"),
            ("Arrival Date:", "arrival_date", "date"),
        ]
        basic_vars = {}
        for i, (label, key, field_type) in enumerate(basic_fields):
            ttk.Label(frame, text=label).grid(row=row, column=0, sticky="e", padx=5, pady=4)
            var = tk.StringVar(value=mouse.get(key, ""))
            basic_vars[key] = var
            if field_type == "strain":
                cb = ttk.Combobox(frame, textvariable=var, values=self.STRAIN_OPTIONS, state="readonly", width=22)
                cb.grid(row=row, column=1, sticky="w", padx=5, pady=4)
                if var.get() not in self.STRAIN_OPTIONS: var.set(self.STRAIN_OPTIONS[0])
            elif field_type == "sex":
                cb = ttk.Combobox(frame, textvariable=var, values=self.SEX_OPTIONS, state="readonly", width=22)
                cb.grid(row=row, column=1, sticky="w", padx=5, pady=4)
                if var.get() not in self.SEX_OPTIONS: var.set(self.SEX_OPTIONS[0])
            elif field_type == "group":
                cb = ttk.Combobox(frame, textvariable=var, values=self.GROUP_OPTIONS, state="readonly", width=22)
                cb.grid(row=row, column=1, sticky="w", padx=5, pady=4)
            elif field_type == "sub_group":
                cb = ttk.Combobox(frame, textvariable=var, values=self.SUB_GROUP_OPTIONS, state="readonly", width=22)
                cb.grid(row=row, column=1, sticky="w", padx=5, pady=4)
            elif field_type == "coat_color":
                cb = ttk.Combobox(frame, textvariable=var, values=self.COAT_COLOR_OPTIONS, state="readonly", width=22)
                cb.grid(row=row, column=1, sticky="w", padx=5, pady=4)
            elif field_type == "date":
                DateEntry(frame, textvariable=var).grid(row=row, column=1, sticky="w", padx=5, pady=4)
            else:
                AutoCorrectEntry(frame, textvariable=var, width=25).grid(row=row, column=1, sticky="w", padx=5, pady=4)
            row += 1

        ttk.Separator(frame, orient="horizontal").grid(row=row, column=0, columnspan=2, sticky="ew", pady=8)
        row += 1
        ttk.Label(frame, text="Experiment Data", font=("", 11, "bold")).grid(
            row=row, column=0, columnspan=2, sticky="w", padx=10, pady=5)
        row += 1

        ttk.Label(frame, text="Date:").grid(row=row, column=0, sticky="e", padx=5, pady=4)
        var_date = tk.StringVar(value=record_date if record_date else datetime.now().strftime("%Y-%m-%d"))
        DateEntry(frame, textvariable=var_date).grid(row=row, column=1, sticky="w", padx=5, pady=4)
        row += 1

        ttk.Label(frame, text="Weight (g):").grid(row=row, column=0, sticky="e", padx=5, pady=4)
        var_weight = tk.StringVar(value=str(record.get("weight", "")) if record and record.get("weight") is not None else "")
        ttk.Entry(frame, textvariable=var_weight, width=15).grid(row=row, column=1, sticky="w", padx=5, pady=4)
        row += 1

        ttk.Label(frame, text="Rotarod Phase:").grid(row=row, column=0, sticky="e", padx=5, pady=4)
        var_rota_phase = tk.StringVar(value=record.get("rotarod_phase", "") if record else "")
        ttk.Combobox(frame, textvariable=var_rota_phase, values=self.PHASE_OPTIONS, state="readonly", width=13).grid(
            row=row, column=1, sticky="w", padx=5, pady=4)
        row += 1

        ttk.Label(frame, text="BB Phase:").grid(row=row, column=0, sticky="e", padx=5, pady=4)
        var_bb_phase = tk.StringVar(value=record.get("bb_phase", "") if record else "")
        ttk.Combobox(frame, textvariable=var_bb_phase, values=self.PHASE_OPTIONS, state="readonly", width=13).grid(
            row=row, column=1, sticky="w", padx=5, pady=4)
        row += 1

        ttk.Label(frame, text="NS:").grid(row=row, column=0, sticky="e", padx=5, pady=4)
        ns_val = record.get("neurological_score") if record else None
        var_ns = tk.StringVar(value=str(ns_val) if ns_val is not None else "")
        ttk.Combobox(frame, textvariable=var_ns, values=self.NS_OPTIONS, state="readonly", width=13).grid(
            row=row, column=1, sticky="w", padx=5, pady=4)
        row += 1

        ttk.Separator(frame, orient="horizontal").grid(row=row, column=0, columnspan=2, sticky="ew", pady=5)
        row += 1

        ttk.Label(frame, text="Grip (10 values):", font=("", 9, "bold")).grid(
            row=row, column=0, columnspan=2, sticky="w", padx=5, pady=3)
        row += 1
        grip = record.get("grip", []) if record else []
        grip_vars = []
        for i in range(10):
            ttk.Label(frame, text=f"{i+1}:").grid(row=row, column=0, sticky="e", padx=5, pady=1)
            var = tk.StringVar(value=str(grip[i]) if i < len(grip) else "")
            ttk.Entry(frame, textvariable=var, width=15).grid(row=row, column=1, sticky="w", padx=5, pady=1)
            grip_vars.append(var)
            row += 1

        ttk.Label(frame, text="Rotarod Trials:", font=("", 9, "bold")).grid(
            row=row, column=0, columnspan=2, sticky="w", padx=5, pady=3)
        row += 1
        rota = record.get("rotarod", []) if record else []
        rota_vars = []
        for i in range(3):
            ttk.Label(frame, text=f"T{i+1} Speed:").grid(row=row, column=0, sticky="e", padx=5, pady=1)
            var_s = tk.StringVar(value=str(rota[i]["speed"]) if i < len(rota) else "")
            ttk.Entry(frame, textvariable=var_s, width=7).grid(row=row, column=1, sticky="w", padx=5, pady=1)
            row += 1
            ttk.Label(frame, text=f"T{i+1} Latency:").grid(row=row, column=0, sticky="e", padx=5, pady=1)
            var_l = tk.StringVar(value=str(rota[i]["latency_to_fall"]) if i < len(rota) else "")
            ttk.Entry(frame, textvariable=var_l, width=7).grid(row=row, column=1, sticky="w", padx=5, pady=1)
            rota_vars.append((var_s, var_l))
            row += 1

        ttk.Label(frame, text="Balance Beam:", font=("", 9, "bold")).grid(
            row=row, column=0, columnspan=2, sticky="w", padx=5, pady=3)
        row += 1
        bb = record.get("balance_beam", {"6mm": [], "12mm": []}) if record else {"6mm": [], "12mm": []}
        bb_vars = {}
        for width in ["6mm", "12mm"]:
            bb_vars[width] = []
            for ti in range(2):
                ttk.Label(frame, text=f"{width} T{ti+1}:").grid(row=row, column=0, sticky="e", padx=5, pady=1)
                var = tk.StringVar(value=str(bb[width][ti]) if ti < len(bb.get(width, [])) else "")
                ttk.Entry(frame, textvariable=var, width=15).grid(row=row, column=1, sticky="w", padx=5, pady=1)
                bb_vars[width].append(var)
                row += 1

        def save_all():
            old_tail_id = mouse.get("tail_id", "")
            new_tail_id = normalize_input(basic_vars["tail_id"].get().strip())
            cage_id = normalize_input(basic_vars["cage_id"].get().strip())
            ear_tag = normalize_input(basic_vars["ear_tag"].get().strip())
            group = basic_vars["group"].get().strip()
            sub_group = basic_vars["sub_group"].get().strip()
            coat_color = basic_vars["coat_color"].get().strip()
            strain = basic_vars["strain"].get().strip()
            sex = basic_vars["sex"].get().strip()
            arrival_date = normalize_input(basic_vars["arrival_date"].get().strip())

            if not new_tail_id:
                messagebox.showwarning("Warning", "Tail ID cannot be empty", parent=dialog)
                return

            success, msg = self.dm.update_mouse(old_tail_id, new_tail_id, cage_id, ear_tag, group, strain, sex, sub_group, coat_color, arrival_date)
            if not success:
                messagebox.showerror("Error", msg, parent=dialog)
                return

            used_tail_id = new_tail_id
            new_date = normalize_input(var_date.get().strip())
            if not new_date:
                messagebox.showwarning("Warning", "Date cannot be empty", parent=dialog)
                return

            w = normalize_input(var_weight.get().strip())
            weight = round(float(w), 2) if w else None
            grip_vals = [float(normalize_input(v.get().strip())) for v in grip_vars if normalize_input(v.get().strip())]
            rota_vals = []
            for vs, vl in rota_vars:
                s = normalize_input(vs.get().strip()); l = normalize_input(vl.get().strip())
                if s and l: rota_vals.append({"speed": float(s), "latency_to_fall": float(l)})
            bb_vals = {"6mm": [], "12mm": []}
            for width in ["6mm", "12mm"]:
                for v in bb_vars[width]:
                    txt = normalize_input(v.get().strip())
                    if txt: bb_vals[width].append(float(txt))
            rp = var_rota_phase.get().strip(); bp = var_bb_phase.get().strip()
            ns_str = var_ns.get().strip()
            ns = int(ns_str) if ns_str else None

            if record_date and record_date != new_date:
                self.dm.delete_record(used_tail_id, record_date)

            has_data = (weight is not None or len(grip_vals) > 0 or len(rota_vals) > 0 or
                        len(bb_vals["6mm"]) > 0 or len(bb_vals["12mm"]) > 0 or ns is not None)

            if has_data:
                success2, msg2 = self.dm.add_record(used_tail_id, new_date, weight, grip_vals, rota_vals, bb_vals, rp, bp, ns)
                if not success2:
                    messagebox.showerror("Error", msg2, parent=dialog)
                    return
            elif record_date:
                self.dm.delete_record(used_tail_id, record_date)

            messagebox.showinfo("Success", "All changes saved.", parent=dialog)
            dialog.destroy()
            self.refresh_all()

        ttk.Separator(frame, orient="horizontal").grid(row=row, column=0, columnspan=2, sticky="ew", pady=8)
        row += 1
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=15)
        ttk.Button(btn_frame, text="💾 Save All", command=save_all).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    def batch_edit_mice(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("Warning", "Please select rows in the table first.")
            return

        cols = list(self.tree["columns"])
        tail_idx = cols.index("Tail ID") if "Tail ID" in cols else 0
        selected_mice = []
        seen = set()
        for item in selected_items:
            vals = self.tree.item(item, "values")
            if vals and len(vals) > tail_idx:
                tid = vals[tail_idx]
                if tid and tid not in seen:
                    selected_mice.append(tid)
                    seen.add(tid)

        if not selected_mice:
            messagebox.showwarning("Warning", "No valid mice found in selection.")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title(f"Batch Edit: {len(selected_mice)} mice selected")
        dialog.geometry("500x500")
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding=15)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text=f"Batch editing {len(selected_mice)} mice:", font=("", 10, "bold")).pack(anchor="w", pady=5)
        ttk.Label(frame, text="(select rows in table, then click Batch Edit)", font=("", 9, "italic")).pack(anchor="w", pady=2)

        mice_text = tk.Text(frame, height=3, width=50, wrap=tk.WORD, font=("", 9))
        mice_text.insert("1.0", ", ".join(selected_mice))
        mice_text.configure(state="disabled")
        mice_text.pack(fill="x", pady=5)

        ttk.Label(frame, text="Only fields you check will be modified:", font=("", 9, "bold"), foreground="blue").pack(anchor="w", pady=5)
        ttk.Separator(frame, orient="horizontal").pack(fill="x", pady=5)

        edit_vars = {}
        for i, (label, key, field_type) in enumerate(self.BATCH_EDITABLE_FIELDS):
            row_frame = ttk.Frame(frame)
            row_frame.pack(fill="x", pady=4)

            modify_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(row_frame, variable=modify_var).pack(side=tk.LEFT, padx=5)
            ttk.Label(row_frame, text=label, width=18, anchor="e").pack(side=tk.LEFT, padx=5)

            value_var = tk.StringVar()
            if field_type == "group":
                ttk.Combobox(row_frame, textvariable=value_var, values=self.GROUP_OPTIONS, state="readonly", width=22).pack(side=tk.LEFT, padx=5)
            elif field_type == "sub_group":
                ttk.Combobox(row_frame, textvariable=value_var, values=self.SUB_GROUP_OPTIONS, state="readonly", width=22).pack(side=tk.LEFT, padx=5)
            elif field_type == "coat_color":
                ttk.Combobox(row_frame, textvariable=value_var, values=self.COAT_COLOR_OPTIONS, state="readonly", width=22).pack(side=tk.LEFT, padx=5)
            elif field_type == "strain":
                ttk.Combobox(row_frame, textvariable=value_var, values=self.STRAIN_OPTIONS, state="readonly", width=22).pack(side=tk.LEFT, padx=5)
            elif field_type == "sex":
                ttk.Combobox(row_frame, textvariable=value_var, values=self.SEX_OPTIONS, state="readonly", width=22).pack(side=tk.LEFT, padx=5)
            else:
                AutoCorrectEntry(row_frame, textvariable=value_var, width=25).pack(side=tk.LEFT, padx=5)

            edit_vars[key] = (modify_var, value_var)

        def apply_batch():
            updates = {}
            for key, (modify_var, value_var) in edit_vars.items():
                if modify_var.get():
                    updates[key] = value_var.get().strip()

            if not updates:
                messagebox.showwarning("Warning", "No fields selected for modification.", parent=dialog)
                return

            msg = f"Apply the following changes to {len(selected_mice)} mice?\n\n"
            for k, v in updates.items():
                label = k
                for l, kk, _ in self.BATCH_EDITABLE_FIELDS:
                    if kk == k:
                        label = l.strip(":")
                        break
                msg += f"  • {label}: {v if v else '(empty)'}\n"

            if not messagebox.askyesno("Confirm Batch Edit", msg, parent=dialog):
                return

            updated_count = self.dm.batch_update_mice(selected_mice, updates)
            messagebox.showinfo("Success", f"Updated {updated_count} mice.", parent=dialog)
            dialog.destroy()
            self.refresh_all()

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=15)
        ttk.Button(btn_frame, text="Apply Changes", command=apply_batch).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    def export_to_excel(self):
        if not HAS_EXCEL:
            messagebox.showerror("Missing Library", "Please install openpyxl:\npip install openpyxl"); return

        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not filepath: return

        filter_group = self.cb_filter_group.get()
        selected_tails = [self.listbox_filter_mice.get(i) for i in self.listbox_filter_mice.curselection()]
        selected_ears = [self.listbox_filter_ear.get(i) for i in self.listbox_filter_ear.curselection()]

        mice_to_export = []
        for m in self.dm.get_all_mice():
            if filter_group and m["group"] != filter_group: continue
            if selected_tails and m.get("tail_id", "") not in selected_tails: continue
            if selected_ears and m.get("ear_tag", "") not in selected_ears: continue
            mice_to_export.append(m)

        if not mice_to_export:
            messagebox.showwarning("Warning", "No data to export.")
            return

        mice_to_export.sort(key=lambda m: natural_sort_key(m.get("tail_id", "")))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "By Mouse (Pivoted)"

        fixed_cols = ["Tail ID", "Cage", "Ear Tag", "Group", "Sub-Group", "Coat Color", "Strain", "Sex", "Arrival Date"]
        for c, col_name in enumerate(fixed_cols, 1):
            cell = ws.cell(row=1, column=c, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        all_dates = set()
        for m in mice_to_export:
            for r in m["records"]:
                all_dates.add(r["date"])
        all_dates = sorted(all_dates)

        metric_cols = [
            ("Wt(g)", "weight"),
            ("GripMax(g)", "grip_max"),
            ("RotaPhase", "rotarod_phase"),
            ("Rotarod(rpm)", "rotarod_speed"),
            ("Rotarod(s)", "rotarod_latency"),
            ("BBPhase", "bb_phase"),
            ("BB6mm(s)", "bb6mm"),
            ("BB12mm(s)", "bb12mm"),
            ("NS", "ns"),
        ]

        col = len(fixed_cols) + 1
        metric_col_map = {}
        for metric_name, metric_key in metric_cols:
            for d in all_dates:
                cell = ws.cell(row=1, column=col, value=f"{metric_name}\n{d}")
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                metric_col_map[(metric_key, d)] = col
                col += 1

        for r_idx, m in enumerate(mice_to_export, 2):
            fixed_values = [
                m.get("tail_id", ""), m.get("cage_id", ""), m.get("ear_tag", ""),
                m.get("group", ""), m.get("sub_group", ""), m.get("coat_color", ""),
                m.get("strain", ""), m.get("sex", ""), m.get("arrival_date", "")
            ]
            for c, val in enumerate(fixed_values, 1):
                ws.cell(row=r_idx, column=c, value=val)

            for (metric_key, d), c_idx in metric_col_map.items():
                val = None
                for rec in m["records"]:
                    if rec["date"] == d:
                        val = self._get_export_metric_value(rec, metric_key)
                        break
                if val is not None:
                    ws.cell(row=r_idx, column=c_idx, value=val)

        for c in range(1, col):
            ws.column_dimensions[get_column_letter(c)].width = 12
        for c in range(1, len(fixed_cols) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 14

        ws.freeze_panes = ws.cell(row=2, column=len(fixed_cols) + 1)

        wb.save(filepath)
        messagebox.showinfo("Success", f"Data exported to:\n{filepath}")

    def _get_export_metric_value(self, record, metric_key):
        if metric_key == "weight":
            return record.get("weight")
        elif metric_key == "grip_max":
            return self.dm.compute_grip_max(record.get("grip", []))
        elif metric_key == "rotarod_phase":
            return record.get("rotarod_phase", "")
        elif metric_key == "rotarod_speed":
            best = self.dm.compute_rotarod_best(record.get("rotarod", []))
            return best["speed"] if best else None
        elif metric_key == "rotarod_latency":
            best = self.dm.compute_rotarod_best(record.get("rotarod", []))
            return best["latency_to_fall"] if best else None
        elif metric_key == "bb_phase":
            return record.get("bb_phase", "")
        elif metric_key == "bb6mm":
            m_val = self.dm.compute_balance_beam_means(record.get("balance_beam", {}))
            return m_val.get("6mm")
        elif metric_key == "bb12mm":
            m_val = self.dm.compute_balance_beam_means(record.get("balance_beam", {}))
            return m_val.get("12mm")
        elif metric_key == "ns":
            return record.get("neurological_score")
        return None

    def delete_selected_mouse(self):
        sel = self.tree.selection()
        if not sel: return
        vals = self.tree.item(sel[0], "values")
        if not vals: return
        cols = list(self.tree["columns"])
        tail_idx = cols.index("Tail ID") if "Tail ID" in cols else 0
        mid = vals[tail_idx]
        if messagebox.askyesno("Confirm", f"Delete mouse {mid} and all its records?"):
            self.dm.delete_mouse(mid); self.refresh_all()

    def delete_selected_record(self):
        sel = self.tree.selection()
        if not sel: return
        vals = self.tree.item(sel[0], "values")
        if not vals or len(vals) < 2: return
        cols = list(self.tree["columns"])
        tail_idx = cols.index("Tail ID") if "Tail ID" in cols else 0
        date_idx = cols.index("Date") if "Date" in cols else 1
        mid = vals[tail_idx]; date = vals[date_idx]
        if date == "No records": return
        if messagebox.askyesno("Confirm", f"Delete record {mid} @ {date}?"):
            self.dm.delete_record(mid, date); self.refresh_all()

    def create_backup(self):
        path = self.dm.backup()
        if path: messagebox.showinfo("Backup", f"Backup created:\n{path}")
        else: messagebox.showwarning("Backup", "No data file to backup yet.")

    # ===================== 图表生成标签页 =====================
    def setup_chart_tab(self):
        control_frame = ttk.Frame(self.tab_chart)
        control_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(control_frame, text="Select Mice (multi-select):").grid(row=0, column=0, sticky="w", padx=5, pady=3)
        self.listbox_mice = tk.Listbox(control_frame, selectmode=tk.MULTIPLE, height=6, width=30, exportselection=False)
        self.listbox_mice.grid(row=0, column=1, rowspan=5, padx=5, pady=3)

        ttk.Label(control_frame, text="Metric:").grid(row=0, column=2, sticky="w", padx=5, pady=3)
        self.var_metric = tk.StringVar(value="Weight")
        metrics = ["Weight", "Grip (Max)", "Rotarod Latency to fall", "Balance Beam 6mm", "Balance Beam 12mm", "Neurological Score"]
        ttk.Combobox(control_frame, textvariable=self.var_metric, values=metrics, state="readonly", width=22).grid(row=0, column=3, sticky="w", padx=5, pady=3)

        ttk.Label(control_frame, text="Chart Type:").grid(row=1, column=2, sticky="w", padx=5, pady=3)
        self.var_chart_type = tk.StringVar(value="line")
        ttk.Radiobutton(control_frame, text="Line Chart", variable=self.var_chart_type, value="line").grid(row=1, column=3, sticky="w")
        ttk.Radiobutton(control_frame, text="Bar Chart", variable=self.var_chart_type, value="bar").grid(row=1, column=4, sticky="w")

        self.var_group_mode = tk.BooleanVar(value=False)
        ttk.Checkbutton(control_frame, text="Group Mean ± SEM", variable=self.var_group_mode).grid(row=2, column=2, columnspan=3, sticky="w", padx=5)

        ttk.Button(control_frame, text="Generate Chart", command=self.generate_chart).grid(row=3, column=2, columnspan=3, pady=10)
        ttk.Button(control_frame, text="Export PNG", command=self.export_chart).grid(row=4, column=2, columnspan=3, pady=5)

        self.chart_frame = ttk.Frame(self.tab_chart)
        self.chart_frame.pack(fill="both", expand=True, padx=10, pady=5)

    def _get_metric_value(self, record, metric):
        if metric == "Weight": return record.get("weight")
        elif metric == "Grip (Max)": return self.dm.compute_grip_max(record.get("grip", []))
        elif metric == "Rotarod Latency to fall":
            best = self.dm.compute_rotarod_best(record.get("rotarod", []))
            return best["latency_to_fall"] if best else None
        elif metric == "Balance Beam 6mm":
            return self.dm.compute_balance_beam_means(record.get("balance_beam", {})).get("6mm")
        elif metric == "Balance Beam 12mm":
            return self.dm.compute_balance_beam_means(record.get("balance_beam", {})).get("12mm")
        elif metric == "Neurological Score": return record.get("neurological_score")
        return None

    def generate_chart(self):
        for w in self.chart_frame.winfo_children(): w.destroy()
        try:
            if self.var_group_mode.get(): self._generate_group_chart()
            else: self._generate_individual_chart()
        except Exception as e:
            messagebox.showerror("Chart Error", f"{e}"); traceback.print_exc()

    def _generate_individual_chart(self):
        selected = [self.listbox_mice.get(i) for i in self.listbox_mice.curselection()]
        if not selected: messagebox.showwarning("Warning", "Select at least one mouse"); return
        metric = self.var_metric.get()
        fig, ax = plt.subplots(figsize=(8, 4.5))
        for mid in selected:
            mouse = self.dm.get_mouse_info(mid)
            if not mouse: continue
            dates, vals = [], []
            for r in mouse["records"]:
                v = self._get_metric_value(r, metric)
                if v is not None: dates.append(r["date"]); vals.append(v)
            if not dates: continue
            label = f"{mid} ({mouse['group']})"
            if self.var_chart_type.get() == "line": ax.plot(dates, vals, marker="o", label=label, linewidth=2)
            else: ax.bar(dates, vals, alpha=0.7, label=label)
        ax.set_xlabel("Date"); ax.set_ylabel(metric); ax.set_title(f"{metric} Over Time"); ax.legend(); ax.grid(True, linestyle="--", alpha=0.5)
        plt.xticks(rotation=45, ha="right"); fig.tight_layout()
        self.current_fig = fig
        FigureCanvasTkAgg(fig, self.chart_frame).draw().get_tk_widget().pack(fill="both", expand=True)

    def _generate_group_chart(self):
        metric = self.var_metric.get()
        groups = {}
        for m in self.dm.get_all_mice():
            g = m["group"]
            if g not in groups: groups[g] = []
            groups[g].append(m)
        if not groups: messagebox.showwarning("Warning", "No data"); return
        all_dates = set()
        for ml in groups.values():
            for m in ml:
                for r in m["records"]:
                    if self._get_metric_value(r, metric) is not None: all_dates.add(r["date"])
        all_dates = sorted(all_dates)
        if not all_dates: messagebox.showwarning("Warning", "No records"); return
        fig, ax = plt.subplots(figsize=(8, 4.5))
        chart_type = self.var_chart_type.get()
        for gn, ml in groups.items():
            means, sems, vd = [], [], []
            for d in all_dates:
                vals = []
                for m in ml:
                    for r in m["records"]:
                        if r["date"] == d:
                            v = self._get_metric_value(r, metric)
                            if v is not None: vals.append(v)
                if len(vals) >= 2:
                    means.append(np.mean(vals)); sems.append(np.std(vals, ddof=1)/np.sqrt(len(vals))); vd.append(d)
                elif len(vals) == 1:
                    means.append(vals[0]); sems.append(0); vd.append(d)
            if vd:
                if chart_type == "line":
                    ax.errorbar(vd, means, yerr=sems, marker="o", capsize=5, label=f"{gn} (n={len(ml)})", linewidth=2)
                else:
                    x = np.arange(len(vd))
                    ax.bar(x, means, yerr=sems, capsize=5, alpha=0.7, label=f"{gn} (n={len(ml)})")
                    ax.set_xticks(x); ax.set_xticklabels(vd, rotation=45, ha="right")
        ax.set_xlabel("Date"); ax.set_ylabel(metric); ax.set_title(f"Group Mean ± SEM — {metric}")
        ax.legend(); ax.grid(True, linestyle="--", alpha=0.5)
        if chart_type == "line": plt.xticks(rotation=45, ha="right")
        fig.tight_layout(); self.current_fig = fig
        FigureCanvasTkAgg(fig, self.chart_frame).draw().get_tk_widget().pack(fill="both", expand=True)

    def export_chart(self):
        if self.current_fig is None: messagebox.showwarning("Warning", "Generate a chart first"); return
        filepath = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG Image", "*.png")])
        if filepath: self.current_fig.savefig(filepath, dpi=150, bbox_inches="tight"); messagebox.showinfo("Success", f"Saved to:\n{filepath}")

    def refresh_all(self):
        tail_ids = self.dm.get_tail_ids()
        ear_tags = self.dm.get_ear_tags()

        self.cb_mouse["values"] = tail_ids
        self.cb_mouse_ear["values"] = ear_tags
        self.cb_regroup_mouse["values"] = tail_ids
        self.cb_regroup_ear_tag["values"] = ear_tags

        if tail_ids: self.cb_mouse.current(0); self.on_mouse_select()

        groups = sorted(set(m["group"] for m in self.dm.get_all_mice()))
        self.cb_filter_group["values"] = [""] + groups

        self.listbox_filter_mice.delete(0, tk.END)
        for mid in tail_ids: self.listbox_filter_mice.insert(tk.END, mid)

        self.listbox_filter_ear.delete(0, tk.END)
        for et in ear_tags: self.listbox_filter_ear.insert(tk.END, et)

        self.listbox_mice.delete(0, tk.END)
        for mid in tail_ids: self.listbox_mice.insert(tk.END, mid)

        self.refresh_view()
        total_mice, total_records = self.dm.get_stats()
        self.status.config(text=f"Data file: {self.dm.filepath} | {total_mice} mice, {total_records} records | Birth: {self.dm.get_birth_date()}")


if __name__ == "__main__":
    root = tk.Tk()
    app = MouseWeightApp(root)
    root.mainloop()
